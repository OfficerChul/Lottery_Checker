import tkinter.font
import requests
from tkinter import *
from bs4 import BeautifulSoup
from tkinter import filedialog
from openpyxl import load_workbook


# 최신 로또 번호 리스트로 return
def get_latest_prize_numbers():
    url = "https://dhlottery.co.kr/common.do?method=main"
    res = requests.get(url)
    res.raise_for_status()
    soup = BeautifulSoup(res.text, "lxml")

    # 당첨번호와 보너스번호
    seq = []
    # 6개의 당첨번호.
    for idx in range(1, 7):
        nums = soup.find("span", attrs={"id": "drwtNo" + str(idx)})
        seq.append(nums.get_text())
    # 보너스 번호
    nums = soup.find("span", attrs={"id": "bnusNo"})
    seq.append(nums.get_text())
    return seq


# 입력받은 로또 번호들 당첨확인.
def check_lotto_numbers(user_xlsx):
    rank = [[], [], [], [], []]
    wb = load_workbook(user_xlsx)
    ws = wb.active
    row_max = ws.max_row
    dat_prize_numbers = [False for i in range(46)]
    for n in prize_numbers:
        dat_prize_numbers[n] = True

    my_list = [[0 for col in range(7)] for row in range(row_max + 1)]
    for x in range(1, row_max + 1):
        for y in range(1, 7):
            n = ws.cell(x, y).value
            my_list[x][y] = n
    func_idx = 1
    for i in my_list[1:]:
        same_point = 0
        check = False
        for j in i:
            if dat_prize_numbers[j]:
                if j == prize_numbers[6]:
                    check = True
                    continue
                same_point += 1
        if same_point == 6:
            rank[0].append(func_idx)
        elif same_point == 5 and check:
            rank[1].append(func_idx)
        elif same_point == 5:
            rank[2].append(func_idx)
        elif same_point == 4:
            rank[3].append(func_idx)
        elif same_point == 3:
            rank[4].append(func_idx)
        func_idx += 1

    results=[]*6
    for i in range(0, 5):
        tmp= str(i+1)+"등: "
        for j in rank[i]:
            tmp += str(j)+"번 "
        # tmp += '\n'
        tmp += f'({len(rank[i])} 개)\n'
        results.append(tmp)

    return results,rank

# 파일 추가및 검사
def add_file():
    txt.delete("1.0", END)


    files = filedialog.askopenfilenames(title="이미지 파일을 선택하세요",
                                        filetypes=(("XLSX 파일", "*.xlsx"), ("모든 파일", "*.*")))
    rst,rank = check_lotto_numbers(files[0])

    for grade in range(0, 5):
        txt.insert(END,rst[grade])
    txt.insert(END,'\n')
    for grade in range(0,5):
        txt.insert(END,str(grade+1)+"등:"+str(len(rank[grade]))+"개 ")


# 다시 하기
def replay():
    txt.delete("1.0", END)


# main 실행
root = Tk()
# 프로그램 제목
root.title("로또 번호 검사 프로그램")

# 화면 크기
root.geometry("700x700+330+100")  # 화면의 가로*세로+x좌표+y좌표. ( x,y좌표는 화면 맨 왼쪽 위 기준)

# font style
font_style = tkinter.font.Font(size=20,weight="bold")

# 오늘의 당첨 번호 화면에 출력
today = Label(root, text="1등 당첨 번호", font=font_style)
today.pack(side="top")
# 로또 6개+ 보너스번호 출력.
prize_numbers = list(map(int, get_latest_prize_numbers()))  # 최신 로또번호 크롤링
seq = Label(root)  # 수열 라밸
seq.pack(side="top")
# [
start = Label(seq, text="[ ", font=font_style)
start.pack(side="left")
for idx in range(0, 6):
    if idx == 5:
        num = Label(seq, text=str(prize_numbers[idx]), font=font_style)
        num.pack(side="left")
    else:
        num = Label(seq, text=str(prize_numbers[idx]) + "  ,", font=font_style)
        num.pack(side="left")
# ]
end = Label(seq, text=" ]", font=font_style)
end.pack(side="left")
# 보너스 번호
bonus_num = Label(seq, text="+ " + str(prize_numbers[6]), font=font_style)
bonus_num.pack(side="left")
# 절취선
line = Label(root, text="ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ"
                        "ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ"
                        "ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ\n\n\n\n\n\n\n")
line.pack(side="top")

# 파일 찾기.

find_file = Button(fg="#FFFFFF", background="#777777",
                   width=7, height=1,
                   text="파일찾기",
                   font=font_style,
                   command=add_file)
find_file.place(x=30, y=120)
txt = Text(root, width=80, height=50,font= 13)
txt.pack(side="top")
# 다시 하기.
again = Button(fg="#FFFFFF", background="#777777",
               width=7, height=1,
               text="지우개",
               font=font_style, command=replay)
again.place(x=170, y=120)
# 루프
root.mainloop()